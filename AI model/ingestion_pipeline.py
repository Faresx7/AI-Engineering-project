import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

from langchain_community.document_loaders import (
    UnstructuredWordDocumentLoader,   
    PyMuPDFLoader,
    TextLoader,
    CSVLoader,
)

from langchain_text_splitters import RecursiveCharacterTextSplitter      # for chunking
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_core.documents import Document
from langchain_chroma import Chroma     # for vector DB

from rank_bm25 import BM25Okapi
from pathlib import Path
import openpyxl
import shutil
import pickle
import json
import re


from retrieval_pipeline import Retrieval


FILE_LOADERS = {
    ".txt": TextLoader,
    ".pdf": PyMuPDFLoader,
    ".docx": UnstructuredWordDocumentLoader,
    ".csv": CSVLoader,
    ".md": TextLoader
    }


class IngestionPipeline:

    def __init__(self,
                 embedding_model_name="sentence-transformers/paraphrase-multilingual-mpnet-base-v2",
                 db_dir=Path(__file__).resolve().parent / "storage" / "db" / "chroma_db",
                 bm25_dir = Path(__file__).resolve().parent / "storage" / "bm25_index.pkl",
                 docs_dir=Path(__file__).resolve().parent / "docs",
                 chunk_overlap=130,
                 chunk_size=800
                 ):
        
        self.doc_dir = Path(docs_dir)
        self.db_dir = Path(db_dir)
        self.bm25_dir = Path(bm25_dir)

        self.embedding_model_name = embedding_model_name
        self.chunk_overlap = chunk_overlap
        self.chunk_size = chunk_size
        self.chunked_files = []
        self.chunks = []


    @staticmethod
    def _excel_loader(file_path: Path) -> list[Document]:
        docs = []
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows or len(rows) < 2:
                continue

            headers = [
                str(h).strip() if h is not None else f"col_{i}"
                for i, h in enumerate(rows[0])
                ]

            for row_idx, row_values in enumerate(rows[1:], start=2):
                row_dict = {
                            headers[i]: str(val).strip()
                            for i, val in enumerate(row_values)
                            if i < len(headers) and val is not None and str(val).strip() != ""
                            }

                if not row_dict:
                        continue

                content = "\n".join([f"{k}: {v}" for k, v in row_dict.items()])

                docs.append(
                    Document(
                            page_content=content,
                            metadata={
                                "source": str(file_path),
                                "sheet": sheet_name,
                                "row": row_idx,
                            },
                        )
                    )

        return docs


    @staticmethod
    def _json_loader(file_path) -> list[Document]:
        
        docs = []
        
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        if isinstance(data, list):
            for item in data:
                if isinstance(item, dict):
                    content = "\n".join([f"{key}: {value}" for key, value in item.items()])
                else:
                    content = str(item)
                
                docs.append(
                    Document(
                        page_content=content,
                        metadata={"source": str(file_path)},
                    )
                )
        else:
            content = json.dumps(data, ensure_ascii=False, indent=2)
            docs.append(
                Document(
                    page_content=content,
                    metadata={"source": str(file_path)},
                )
            )
        
        return docs


    def _load_documents(self, allowed_extensions: list[str] | None = None):

        
        if not self.doc_dir.exists():
            raise FileNotFoundError(f'Path "{self.doc_dir}" does not exist')

        all_supported_ext = set(FILE_LOADERS.keys()) | {".json"} | {".xlsx", ".xls"}

        if allowed_extensions:
            user_ext = {ext.lower() if ext.startswith(".") else f".{ext}" for ext in allowed_extensions}
            valid_ext = user_ext & all_supported_ext  # Only keep extensions that are actually supported
        else:
            valid_ext = all_supported_ext

        docs = []
        failed_files = []
        
        for file_path in self.doc_dir.rglob("*"):
            if file_path.is_file():
                extension = file_path.suffix.lower()

                if extension in valid_ext:
                    try:
                        if extension == ".json":
                            json_docs = self._json_loader(file_path)
                            docs.extend(json_docs)


                        elif extension in [".xlsx", ".xls"]:
                            excel_docs = self._excel_loader(file_path)
                            docs.extend(excel_docs)

                            
                        elif extension in FILE_LOADERS:
                            
                            loader_class = FILE_LOADERS[extension]
                            if extension in [".txt", ".md", ".csv"]:
                                loader = loader_class(str(file_path), encoding="utf-8")
                            else:
                                loader = loader_class(str(file_path))

                            docs.extend(loader.load())

                    except Exception as e:
                        failed_files.append(file_path.name)
                        print(
                            f"⚠️ Warning: Failed to load {file_path.name}. "
                            f"Error: {e}"
                        )

        if not docs:
            if failed_files:
                raise RuntimeError(
                    f'Files were found in "{self.doc_dir}", but none could be loaded. '
                    "For PDFs, install the PDF dependency with "
                    "`pip install pymupdf`. Scanned/image-only PDFs need OCR because "
                    "PyMuPDFLoader can only extract embedded text."
                )
            raise FileNotFoundError(
                f'No supported files found in "{self.doc_dir}" folder'
            )

        # A PDF may open successfully but contain no embedded text (for example,
        # a scanned document). Such a file cannot be indexed by this pipeline.
        empty_pdf_files = {
            Path(doc.metadata.get("source", "")).name
            for doc in docs
            if doc.metadata.get("source", "").lower().endswith(".pdf")
            and not doc.page_content.strip()
        }
        if empty_pdf_files:
            print(
                "⚠️ Warning: These PDFs contain no extractable text and may need OCR: "
                + ", ".join(sorted(empty_pdf_files)) +
                ".\nTry to use PDF files that not scanned."
            )
            
            docs = [
                doc for doc in docs
                if Path(doc.metadata.get("source", "")).name not in empty_pdf_files
            ]

        return docs


    def _chunk_documents(self,docs):
        """
        Splits documents into smaller chunks for embedding and retrieval.
        
        JSON files are kept as whole documents to preserve their structure.
        Other document types (PDF, TXT, etc.) are split using RecursiveCharacterTextSplitter
        with configurable chunk size and overlap to maintain context between chunks.
        This ensures text documents are not too large for embedding models while
        preserving semantic meaning through chunk overlap.
        """
        
        final_chunks = []
        text_docs = []

        text_splitter = RecursiveCharacterTextSplitter(chunk_size = self.chunk_size,
                                                    chunk_overlap = self.chunk_overlap,     # helps model keep the meaning
                                                    separators=["\n\n","\n",". ", " ", ""]
                                                    )


        for doc in docs:
            source = str(doc.metadata.get("source", "")).lower()
            if source.endswith((".json", ".csv", ".xlsx", ".xls")):
                final_chunks.append(doc)
                self.chunked_files.append(doc.metadata['source'])
            else:
                text_docs.append(doc)
                self.chunked_files.append(doc.metadata['source'])

        if text_docs:
            text_chunks = text_splitter.split_documents(text_docs)
            final_chunks.extend(text_chunks)

        return final_chunks


    def _create_vector_store(self, chunks):
        """
        Embeds text chunks using HuggingFace model and stores them in Chroma DB.
        Model max context capacity: 512 tokens (~1000-1200 chars).
        """
        
        if self.db_dir.exists():
            try:
                shutil.rmtree(self.db_dir)
            except Exception as e:
                print(f"⚠️ Warning: Could not clean old DB directory fully: {e}")

        embedding_model = HuggingFaceEmbeddings(model_name=self.embedding_model_name)
        # ``client`` is provided at runtime by HuggingFaceEmbeddings but is
        # not exposed in its static type definition.
        embedding_client = getattr(embedding_model, "client", None)
        if embedding_client is not None:
            embedding_client.max_seq_length = 512


        vector_store = Chroma.from_documents(
            documents= chunks,
            embedding = embedding_model,
            persist_directory = str(self.db_dir),
            collection_metadata= {"hnsw:space":"cosine"}
        )
        
        return vector_store


    def preview_chunks(self, chunks=None, limit=3):
        """limit = -1 or None: prints all chunks limit = N: prints first N chunks"""
        chunks = self.chunks if chunks is None else chunks

        if not chunks:
            print("⚠️ No chunks available to preview!")
            return

        
        show_all = limit is None or limit < 0
        target_chunks = chunks if show_all else chunks[:limit]
        header_count = len(chunks) if show_all else min(limit, len(chunks))

        print(f"\n🔍 --- Previewing ({header_count} / {len(chunks)}) Chunks ---")

        for i, chunk in enumerate(target_chunks, 1):
            print(f"\n📦 Chunk #{i}:")
            print(f"📄 Source: {chunk.metadata.get('source', 'N/A')}")
            print(
                f"📏 Length: {len(chunk.page_content)} chars |"
                f" {len(chunk.page_content.split())} words"
            )
            print("📝 Content:")
            print("-" * 50)
            print(chunk.page_content)
            print("-" * 50)


    def _create_bm25_index(self,chunks):
        """Tokenize document chunks and save BM25 index for keyword-based retrieval.
        
        Args:
            chunks: List of document chunks to tokenize
        """
        raw_text = [doc.page_content for doc in chunks]

        tokenized_chunks = [Retrieval.tokenize_and_remove_stopwords(doc) for doc in raw_text]
        bm25 = BM25Okapi(tokenized_chunks)

        self.bm25_dir.parent.mkdir(parents=True, exist_ok=True)

        with open(self.bm25_dir,"wb") as file:
            pickle.dump({"raw_text":raw_text, "bm25":bm25}, file)


    def run(self, verbose=True, preview_n_chunks=0, allowed_extensions: list[str] | None = None):

        if verbose:
            print("🚀 [1/4] Loading documents...")
        docs = self._load_documents(allowed_extensions = allowed_extensions)

        if verbose:
            print(f"✂️ [2/4] Chunking {len(docs)} document(s)...")
        chunks = self._chunk_documents(docs)
        self.chunks = chunks
        
        if preview_n_chunks != 0:
            self.preview_chunks(chunks,limit = preview_n_chunks)

        if verbose:
            print(f"💾 [3/4] Embedding & saving {len(chunks)} chunks to Chroma DB...")
        vector_store = self._create_vector_store(chunks)

        if verbose:
            print("💿 [4/4] creating bm25 indexing...")
        self._create_bm25_index(self.chunks)


        if verbose:
            print("✅ Ingestion Pipeline completed successfully!")
            print(f"chunked files:\n{set(m.group() for path in self.chunked_files if (m := re.search(r'[^/\\]+$', path)))}")


        return vector_store


if __name__ == "__main__":
    pipeline = IngestionPipeline()

    vdb = pipeline.run(preview_n_chunks=60)
